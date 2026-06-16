"""Data export service for trend and alarm data."""
import io
import csv
import time
import random
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse

from app.models.schemas import ExportConfig, TrendDataPoint, Alarm, ExportTask, ExportHistoryRecord

MOCK_DEVICES = [
    {"id": "dev1", "name": "温湿度传感器-A区", "ip": "192.168.1.101", "port": 502, "slave_id": 1, "online": True,
     "registers": [
         {"address": 0, "name": "温度", "type": "holding", "unit": "°C", "base_value": 25.6},
         {"address": 1, "name": "湿度", "type": "holding", "unit": "%RH", "base_value": 62.3},
         {"address": 2, "name": "露点", "type": "holding", "unit": "°C", "base_value": 17.8},
     ]},
    {"id": "dev2", "name": "压力变送器-B区", "ip": "192.168.1.102", "port": 502, "slave_id": 2, "online": True,
     "registers": [
         {"address": 0, "name": "管道压力", "type": "holding", "unit": "MPa", "base_value": 3.45},
         {"address": 1, "name": "差压", "type": "holding", "unit": "kPa", "base_value": 0.12},
     ]},
    {"id": "dev3", "name": "电机控制器-C区", "ip": "192.168.1.103", "port": 502, "slave_id": 3, "online": False,
     "registers": [
         {"address": 0, "name": "转速", "type": "holding", "unit": "RPM", "base_value": 1480},
         {"address": 1, "name": "电流", "type": "holding", "unit": "A", "base_value": 12.5},
     ]},
    {"id": "dev4", "name": "流量计-D区", "ip": "192.168.1.104", "port": 502, "slave_id": 4, "online": True,
     "registers": [
         {"address": 0, "name": "瞬时流量", "type": "holding", "unit": "L/min", "base_value": 156.7},
         {"address": 1, "name": "累计流量", "type": "holding", "unit": "L", "base_value": 98234},
     ]},
]

export_history: List[ExportHistoryRecord] = []
export_tasks: List[ExportTask] = []


def _generate_mock_trend_data(config: ExportConfig) -> List[TrendDataPoint]:
    """Generate mock trend data for export."""
    data: List[TrendDataPoint] = []
    device_map = {d["id"]: d for d in MOCK_DEVICES}

    interval_ms = 1000
    current_time = config.start_time

    while current_time <= config.end_time:
        for device_id in config.device_ids:
            device = device_map.get(device_id)
            if not device:
                continue

            for reg in device["registers"]:
                if config.register_names and reg["name"] not in config.register_names:
                    continue

                noise = random.uniform(-0.02, 0.02) * reg["base_value"]
                value = round(reg["base_value"] + noise, 2)

                data.append(TrendDataPoint(
                    time=current_time,
                    device_id=device_id,
                    device_name=device["name"],
                    register=str(reg["address"]),
                    register_name=reg["name"],
                    value=value,
                    unit=reg["unit"]
                ))

        current_time += interval_ms

    return data


def _generate_mock_alarm_data(config: ExportConfig) -> List[Alarm]:
    """Generate mock alarm data for export."""
    device_map = {d["id"]: d for d in MOCK_DEVICES}
    level_map = {0: "info", 1: "warning", 2: "critical"}
    message_templates = {
        "温度": "温度超限: {value}°C",
        "压力": "压力异常: {value}MPa",
        "电流": "电流过高: {value}A",
        "湿度": "湿度超标: {value}%RH",
    }

    alarms: List[Alarm] = []
    duration = config.end_time - config.start_time
    alarm_count = max(1, int(duration / 60000))

    for i in range(alarm_count):
        device_id = random.choice(config.device_ids)
        device = device_map.get(device_id)
        if not device:
            continue

        reg = random.choice(device["registers"])
        level_idx = random.choice([0, 1, 2])
        level = level_map[level_idx]

        if config.alarm_levels and level not in config.alarm_levels:
            continue

        timestamp = config.start_time + random.uniform(0, duration)
        template = message_templates.get(reg["name"], "{name}异常: {value}{unit}")
        value = round(reg["base_value"] * random.uniform(1.1, 1.5), 2)
        message = template.format(name=reg["name"], value=value, unit=reg["unit"])

        alarms.append(Alarm(
            id=f"alarm_{int(timestamp * 1000)}_{i}",
            device_id=device_id,
            register=reg["name"],
            message=f"{device['name']} {message}",
            level=level,
            timestamp=timestamp,
            acknowledged=random.choice([True, False])
        ))

    alarms.sort(key=lambda a: a.timestamp, reverse=True)
    return alarms


def _format_time(timestamp: float) -> str:
    return datetime.fromtimestamp(timestamp / 1000).strftime("%Y-%m-%d %H:%M:%S")


def _format_file_size(bytes_size: int) -> str:
    if bytes_size < 1024:
        return f"{bytes_size} B"
    if bytes_size < 1024 * 1024:
        return f"{bytes_size / 1024:.1f} KB"
    return f"{bytes_size / (1024 * 1024):.1f} MB"


def _generate_filename(config: ExportConfig) -> str:
    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    type_name = "趋势数据" if config.export_type == "trend" else "告警明细"
    return f"{type_name}_{timestamp}.{config.format}"


def export_to_csv_trend(data: List[TrendDataPoint]) -> io.BytesIO:
    """Export trend data to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)

    headers = ["时间", "设备ID", "设备名称", "寄存器地址", "指标名称", "数值", "单位"]
    writer.writerow(headers)

    level_map_cn = {"info": "信息", "warning": "警告", "critical": "严重"}

    for d in data:
        writer.writerow([
            _format_time(d.time),
            d.device_id,
            d.device_name,
            d.register,
            d.register_name,
            d.value,
            d.unit
        ])

    output.seek(0)
    return io.BytesIO(output.getvalue().encode("utf-8-sig"))


def export_to_csv_alarm(alarms: List[Alarm]) -> io.BytesIO:
    """Export alarm data to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)

    device_map = {d["id"]: d["name"] for d in MOCK_DEVICES}
    level_map_cn = {"info": "信息", "warning": "警告", "critical": "严重"}

    headers = ["时间", "设备ID", "设备名称", "指标", "告警信息", "级别", "是否确认"]
    writer.writerow(headers)

    for a in alarms:
        writer.writerow([
            _format_time(a.timestamp),
            a.device_id,
            device_map.get(a.device_id, ""),
            a.register,
            a.message,
            level_map_cn.get(a.level, a.level),
            "是" if a.acknowledged else "否"
        ])

    output.seek(0)
    return io.BytesIO(output.getvalue().encode("utf-8-sig"))


def export_to_xlsx_trend(data: List[TrendDataPoint]) -> io.BytesIO:
    """Export trend data to XLSX format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "趋势数据"

    headers = ["时间", "设备ID", "设备名称", "寄存器地址", "指标名称", "数值", "单位"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, d in enumerate(data, 2):
        ws.cell(row=row, column=1, value=_format_time(d.time))
        ws.cell(row=row, column=2, value=d.device_id)
        ws.cell(row=row, column=3, value=d.device_name)
        ws.cell(row=row, column=4, value=d.register)
        ws.cell(row=row, column=5, value=d.register_name)
        ws.cell(row=row, column=6, value=d.value)
        ws.cell(row=row, column=7, value=d.unit)

    column_widths = [20, 12, 22, 14, 14, 12, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_xlsx_alarm(alarms: List[Alarm]) -> io.BytesIO:
    """Export alarm data to XLSX format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "告警明细"

    device_map = {d["id"]: d["name"] for d in MOCK_DEVICES}
    level_map_cn = {"info": "信息", "warning": "警告", "critical": "严重"}
    level_fills = {
        "info": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "warning": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "critical": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    }

    headers = ["时间", "设备ID", "设备名称", "指标", "告警信息", "级别", "是否确认"]
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, a in enumerate(alarms, 2):
        ws.cell(row=row, column=1, value=_format_time(a.timestamp))
        ws.cell(row=row, column=2, value=a.device_id)
        ws.cell(row=row, column=3, value=device_map.get(a.device_id, ""))
        ws.cell(row=row, column=4, value=a.register)
        ws.cell(row=row, column=5, value=a.message)
        level_cell = ws.cell(row=row, column=6, value=level_map_cn.get(a.level, a.level))
        if a.level in level_fills:
            level_cell.fill = level_fills[a.level]
        ws.cell(row=row, column=7, value="是" if a.acknowledged else "否")

    column_widths = [20, 12, 22, 12, 35, 10, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_data(config: ExportConfig) -> StreamingResponse:
    """Export data based on configuration."""
    filename = _generate_filename(config)

    if config.export_type == "trend":
        data = _generate_mock_trend_data(config)
        if config.format == "xlsx":
            content = export_to_xlsx_trend(data)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            content = export_to_csv_trend(data)
            media_type = "text/csv; charset=utf-8"
    else:
        alarms = _generate_mock_alarm_data(config)
        if config.format == "xlsx":
            content = export_to_xlsx_alarm(alarms)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            content = export_to_csv_alarm(alarms)
            media_type = "text/csv; charset=utf-8"

    file_size = len(content.getvalue())

    history_record = ExportHistoryRecord(
        id=f"hist_{int(time.time() * 1000)}",
        name=filename,
        type=config.export_type,
        format=config.format,
        device_count=len(config.device_ids),
        time_range=f"{_format_time(config.start_time)} ~ {_format_time(config.end_time)}",
        created_at=time.time() * 1000,
        file_size=_format_file_size(file_size)
    )
    export_history.insert(0, history_record)
    if len(export_history) > 20:
        export_history.pop()

    response = StreamingResponse(
        iter([content.getvalue()]),
        media_type=media_type
    )
    response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{filename}"
    return response


def get_export_history() -> List[ExportHistoryRecord]:
    """Get export history records."""
    return export_history


def get_register_names() -> List[str]:
    """Get all available register names."""
    names = set()
    for device in MOCK_DEVICES:
        for reg in device["registers"]:
            names.add(reg["name"])
    return list(names)
